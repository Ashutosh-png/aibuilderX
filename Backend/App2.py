from flask import Flask, request, jsonify
from flask_cors import CORS
from together import Together
import PyPDF2
import docx
import openpyxl  # Changed from xlrd to openpyxl
 
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB
CORS(app)
 
# Initialize the Together client
#client = Together(api_key="84853560af1fe58e69906ac38f347ee8136630b0e34067d2cbcfd36f8d0aec34")  # Make sure to keep your API key secure
 
@app.route('/generate', methods=['POST'])
def generate():
    # Get the prompt and additional parameters from the request
    prompt = request.json.get('prompt')
    summary_type = request.json.get('summaryType')
    summary_length = request.json.get('summaryLength')
    formatting = request.json.get('formatting')
 
    if not prompt:
        return jsonify({'error': 'Prompt is required'}), 400
 
    prompt2 = f"summarize the below in {summary_type} format, in {summary_length} length using {formatting}. The prompt is: {prompt}"
 
    # Create a new chat completion
    client = Groq(api_key="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")
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": prompt2
                }
            ],
            model="llama3-8b-8192",
        )
        response = chat_completion.choices[0].message.content
 
    # Return the response as JSON
    return jsonify({'response': response})
 
 
@app.route('/upload', methods=['POST'])
def upload():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'No file provided'}), 400
       
        text = ''
        if file.filename.endswith('.pdf'):
            pdf_file = PyPDF2.PdfReader(file)
            for page in pdf_file.pages:
                text += page.extract_text() or ''
        elif file.filename.endswith('.docx'):
            doc = docx.Document(file)
            for para in doc.paragraphs:
                text += para.text
        elif file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file)
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += ' '.join(map(str, row)) + ' '
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
 
        # Create a new chat completion
        client = Groq(api_key="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")
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": prompt2
                }
            ],
            model="llama3-8b-8192",
        )
        response = chat_completion.choices[0].message.content
 
        return jsonify({'response': response})
 
    except Exception as e:
        return jsonify({'error': str(e)}), 500
 
if __name__ == '__main__':
    app.run(debug=True)
 